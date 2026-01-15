#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
拆分Excel测试用例单元格脚本
用于split-cells skill
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_column_by_keyword(ws, keywords, start_row=1):
    """
    根据关键词查找列索引
    
    Args:
        ws: worksheet对象
        keywords: 关键词列表，如['用例名称', '用例名', '名称']
        start_row: 开始查找的行（默认第1行，即表头）
    
    Returns:
        列索引（1-based），如果找不到返回None
    """
    if start_row > ws.max_row:
        return None
    
    # 读取表头行
    header_row = start_row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            cell_value = str(cell.value).strip()
            for keyword in keywords:
                if keyword in cell_value:
                    return col_idx
    return None


def get_cell_value(ws, row, col):
    """
    安全获取单元格值
    
    Args:
        ws: worksheet对象
        row: 行号（1-based）
        col: 列号（1-based）
    
    Returns:
        单元格值，如果单元格不存在或为空返回None
    """
    if row > ws.max_row or col > ws.max_column:
        return None
    
    cell = ws.cell(row=row, column=col)
    value = cell.value
    if value is None:
        return None
    
    value_str = str(value).strip()
    return value_str if value_str else None


def split_by_hash(content):
    """
    按#拆分内容，每个#及后续内容（直到下一个#之前或到末尾）拆分成一个片段
    
    例如: "#步骤1#步骤2#步骤3" -> ["#步骤1", "#步骤2", "#步骤3"]
    
    Args:
        content: 要拆分的内容
    
    Returns:
        拆分后的列表，每个元素包含#号
    """
    if content is None:
        return []
    
    content_str = str(content).strip()
    if not content_str:
        return []
    
    # 找到所有#的位置
    parts = []
    start_idx = 0
    
    while True:
        hash_idx = content_str.find('#', start_idx)
        if hash_idx == -1:
            break
        
        # 找到下一个#的位置，或者到末尾
        next_hash_idx = content_str.find('#', hash_idx + 1)
        if next_hash_idx == -1:
            # 到末尾，包含当前#到字符串末尾
            parts.append(content_str[hash_idx:])
            break
        else:
            # 到下一个#之前，包含当前#到下一个#之前
            parts.append(content_str[hash_idx:next_hash_idx])
            start_idx = next_hash_idx
    
    return parts


def find_module_columns(ws, start_row=1):
    """
    查找所有X级模块列
    
    Args:
        ws: worksheet对象
        start_row: 表头行
    
    Returns:
        列索引列表
    """
    module_cols = []
    if start_row > ws.max_row:
        return module_cols
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=col_idx)
        if cell.value:
            cell_value = str(cell.value).strip()
            # 匹配"X级模块"格式
            if '级模块' in cell_value:
                module_cols.append(col_idx)
    
    return module_cols


def split_cells(input_file, output_file=None):
    """
    拆分Excel测试用例单元格
    
    Args:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径（如果为None，则覆盖原文件）
    """
    if output_file is None:
        output_file = input_file
    
    # 加载工作簿
    wb = load_workbook(input_file, data_only=True, keep_vba=False)
    
    # 获取第一个sheet
    ws = wb.active
    
    print(f"处理Sheet: {ws.title}")
    print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")
    
    # 查找列索引
    step_desc_col = find_column_by_keyword(ws, ['步骤描述', '步骤', '描述'], 1)
    expected_result_col = find_column_by_keyword(ws, ['预期结果', '结果', '预期'], 1)
    case_name_col = find_column_by_keyword(ws, ['用例名称', '用例名', '名称'], 1)
    precondition_col = find_column_by_keyword(ws, ['前置条件', '前置'], 1)
    
    if step_desc_col is None:
        print("错误: 找不到'步骤描述'列")
        sys.exit(1)
    if expected_result_col is None:
        print("错误: 找不到'预期结果'列")
        sys.exit(1)
    if case_name_col is None:
        print("错误: 找不到'用例名称'列")
        sys.exit(1)
    
    # 查找所有X级模块列
    module_cols = find_module_columns(ws, 1)
    print(f"找到 {len(module_cols)} 个模块列: {[get_column_letter(col) for col in module_cols]}")
    
    print(f"步骤描述列: {get_column_letter(step_desc_col)} (列{step_desc_col})")
    print(f"预期结果列: {get_column_letter(expected_result_col)} (列{expected_result_col})")
    print(f"用例名称列: {get_column_letter(case_name_col)} (列{case_name_col})")
    if precondition_col:
        print(f"前置条件列: {get_column_letter(precondition_col)} (列{precondition_col})")
    
    # 第一步：拆分步骤描述和预期结果列
    print("\n步骤1: 拆分步骤描述和预期结果列...")
    
    # 先收集所有需要拆分的行数据
    rows_to_split = []  # 存储(row_index, step_parts, result_parts, row_data)
    
    for row in range(2, ws.max_row + 1):  # 从第2行开始（第1行是表头）
        step_value = get_cell_value(ws, row, step_desc_col)
        result_value = get_cell_value(ws, row, expected_result_col)
        
        # 拆分步骤描述
        step_parts = split_by_hash(step_value) if step_value else []
        # 拆分预期结果
        result_parts = split_by_hash(result_value) if result_value else []
        
        # 确定需要拆分的数量（取两者的最大值）
        max_parts = max(len(step_parts), len(result_parts))
        
        if max_parts > 1:
            # 读取当前行的所有列数据
            row_data = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data[col] = cell_value
            
            rows_to_split.append((row, step_parts, result_parts, row_data, max_parts))
    
    # 从后往前插入新行，避免索引变化
    total_new_rows = 0
    for row, step_parts, result_parts, row_data, max_parts in reversed(rows_to_split):
        # 更新第一行（原行）
        if len(step_parts) > 0:
            ws.cell(row=row, column=step_desc_col).value = step_parts[0]
        if len(result_parts) > 0:
            ws.cell(row=row, column=expected_result_col).value = result_parts[0]
        
        # 插入新行
        for i in range(1, max_parts):
            ws.insert_rows(row + 1)
            # 复制所有列的数据
            for col, value in row_data.items():
                ws.cell(row=row + 1, column=col).value = value
            # 设置步骤描述
            if i < len(step_parts):
                ws.cell(row=row + 1, column=step_desc_col).value = step_parts[i]
            else:
                ws.cell(row=row + 1, column=step_desc_col).value = None
            # 设置预期结果
            if i < len(result_parts):
                ws.cell(row=row + 1, column=expected_result_col).value = result_parts[i]
            else:
                ws.cell(row=row + 1, column=expected_result_col).value = None
            total_new_rows += 1
    
    print(f"拆分了 {len(rows_to_split)} 行，共插入 {total_new_rows} 行新数据")
    
    # 第二步：清空步骤描述和预期结果两列中全部单元格里的#
    print("\n步骤2: 清空步骤描述和预期结果列中的#号...")
    cleared_count = 0
    for row in range(2, ws.max_row + 1):
        # 处理步骤描述列
        step_value = ws.cell(row=row, column=step_desc_col).value
        if step_value is not None:
            step_str = str(step_value)
            if step_str.startswith('#'):
                new_step_value = step_str[1:] if len(step_str) > 1 else ''
                ws.cell(row=row, column=step_desc_col).value = new_step_value
                cleared_count += 1
        
        # 处理预期结果列
        result_value = ws.cell(row=row, column=expected_result_col).value
        if result_value is not None:
            result_str = str(result_value)
            if result_str.startswith('#'):
                new_result_value = result_str[1:] if len(result_str) > 1 else ''
                ws.cell(row=row, column=expected_result_col).value = new_result_value
                cleared_count += 1
    
    print(f"清空了 {cleared_count} 个单元格中的#号")
    
    # 第三步：合并X级模块、用例名称、前置条件列中相同内容的单元格
    print("\n步骤3: 合并相同内容的单元格...")
    
    # 需要合并的列
    merge_cols = module_cols + [case_name_col]
    if precondition_col:
        merge_cols.append(precondition_col)
    
    print(f"需要合并的列: {[get_column_letter(col) for col in merge_cols]}")
    
    # 按列处理合并
    for col in merge_cols:
        merged_count = 0
        start_row = None
        last_value = None
        
        for row in range(2, ws.max_row + 1):
            current_value = ws.cell(row=row, column=col).value
            
            # 标准化值（转换为字符串进行比较）
            current_str = str(current_value).strip() if current_value is not None else ''
            last_str = str(last_value).strip() if last_value is not None else ''
            
            if current_str == last_str and current_str:
                # 值相同，继续合并范围
                if start_row is None:
                    start_row = row - 1
            else:
                # 值不同，合并之前的范围
                if start_row is not None and row - 1 > start_row:
                    ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{row - 1}')
                    merged_count += 1
                    start_row = None
                last_value = current_value
        
        # 处理最后一段
        if start_row is not None and ws.max_row > start_row:
            ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{ws.max_row}')
            merged_count += 1
        
        print(f"  列{get_column_letter(col)}: 合并了 {merged_count} 组单元格")
    
    # 第四步：设置所有行的高度为50磅
    print("\n步骤4: 设置所有行的高度为50磅...")
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 50
    print(f"已设置 {ws.max_row} 行的行高为50磅")
    
    # 第五步：删除尾部的空行
    print("\n步骤5: 删除尾部的空行...")
    initial_max_row = ws.max_row
    deleted_count = 0
    
    # 从后往前遍历，找到第一个非空行
    last_non_empty_row = None
    for row in range(ws.max_row, 0, -1):
        # 检查整行是否为空
        is_empty = True
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                # 检查是否为空字符串或只包含空白字符
                if isinstance(cell_value, str):
                    if cell_value.strip():
                        is_empty = False
                        break
                else:
                    is_empty = False
                    break
        
        if not is_empty:
            last_non_empty_row = row
            break
    
    # 删除last_non_empty_row之后的所有空行
    if last_non_empty_row is not None and last_non_empty_row < ws.max_row:
        rows_to_delete = list(range(last_non_empty_row + 1, ws.max_row + 1))
        # 从后往前删除，避免索引变化
        for row in reversed(rows_to_delete):
            ws.delete_rows(row)
            deleted_count += 1
        print(f"删除了 {deleted_count} 个尾部空行（从行{last_non_empty_row + 1}到行{initial_max_row}）")
    else:
        print("未发现尾部空行")
    
    # 保存文件
    try:
        wb.save(output_file)
        print(f"\n✅ 处理完成！文件已保存到: {output_file}")
    except Exception as e:
        print(f"\n❌ 保存文件时出错: {e}")
        sys.exit(1)
    finally:
        wb.close()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python split_cells.py <input_file> [<output_file>]")
        print("  input_file: 输入Excel文件路径")
        print("  output_file: 输出Excel文件路径（可选，默认覆盖原文件）")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    split_cells(input_file, output_file)
