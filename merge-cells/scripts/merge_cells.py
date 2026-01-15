#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
合并Excel测试用例单元格脚本
用于merge-cells skill
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_column_by_keyword(ws, keywords, start_row=1):
    """
    根据关键词查找列索引
    
    Args:
        ws: worksheet对象
        keywords: 关键词列表，如['用例名称', '用例名', '名称']
        start_row: 开始查找的行（默认第1行，即表头）
    
    Returns:
        列索引（1-based），如果找不到返回None
    """
    if start_row > ws.max_row:
        return None
    
    # 读取表头行
    header_row = start_row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            cell_value = str(cell.value).strip()
            for keyword in keywords:
                if keyword in cell_value:
                    return col_idx
    return None


def add_hash_prefix(value):
    """
    给单元格内容前加上#号（如果还没有）
    如果单元格为空，也返回#
    
    Args:
        value: 单元格值
    
    Returns:
        处理后的值
    """
    if value is None:
        return '#'
    
    value_str = str(value).strip()
    if not value_str:
        return '#'
    
    if not value_str.startswith('#'):
        return '#' + value_str
    return value_str


def get_cell_value(ws, row, col):
    """
    安全获取单元格值
    
    Args:
        ws: worksheet对象
        row: 行号（1-based）
        col: 列号（1-based）
    
    Returns:
        单元格值，如果单元格不存在或为空返回None
    """
    if row > ws.max_row or col > ws.max_column:
        return None
    
    cell = ws.cell(row=row, column=col)
    value = cell.value
    if value is None:
        return None
    
    value_str = str(value).strip()
    return value_str if value_str else None


def merge_cells(input_file, output_file=None):
    """
    合并Excel测试用例单元格
    
    Args:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径（如果为None，则覆盖原文件）
    """
    if output_file is None:
        output_file = input_file
    
    # 加载工作簿
    wb = load_workbook(input_file, data_only=True, keep_vba=False)
    
    # 获取第一个sheet
    ws = wb.active
    
    print(f"处理Sheet: {ws.title}")
    print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")
    
    # 查找列索引
    case_name_col = find_column_by_keyword(ws, ['用例名称', '用例名', '名称'], 1)
    step_desc_col = find_column_by_keyword(ws, ['步骤描述', '步骤', '描述'], 1)
    expected_result_col = find_column_by_keyword(ws, ['预期结果', '结果', '预期'], 1)
    
    if case_name_col is None:
        print("错误: 找不到'用例名称'列")
        sys.exit(1)
    if step_desc_col is None:
        print("错误: 找不到'步骤描述'列")
        sys.exit(1)
    if expected_result_col is None:
        print("错误: 找不到'预期结果'列")
        sys.exit(1)
    
    print(f"用例名称列: {get_column_letter(case_name_col)} (列{case_name_col})")
    print(f"步骤描述列: {get_column_letter(step_desc_col)} (列{step_desc_col})")
    print(f"预期结果列: {get_column_letter(expected_result_col)} (列{expected_result_col})")
    
    # 先判断整个文档的最后一行（用例名称列有值的最后一行）
    last_row_with_case_name = None
    for row in range(ws.max_row, 1, -1):
        case_name_value = get_cell_value(ws, row, case_name_col)
        if case_name_value:
            last_row_with_case_name = row
            break
    
    if last_row_with_case_name is None:
        print("警告: 用例名称列没有找到有值的行")
        last_row_with_case_name = ws.max_row
    
    print(f"用例名称列最后有值的行: {last_row_with_case_name}")
    
    # 第一步：给步骤描述和预期结果列的每个单元格内容前加上#号
    print("\n步骤1: 给步骤描述和预期结果列添加#号前缀...")
    modified_count = 0
    for row in range(2, ws.max_row + 1):  # 从第2行开始（第1行是表头）
        # 处理步骤描述列（包括空单元格）
        step_value = get_cell_value(ws, row, step_desc_col)
        new_step_value = add_hash_prefix(step_value)
        # 如果单元格为空或值发生变化，则更新
        if step_value is None or new_step_value != step_value:
            ws.cell(row=row, column=step_desc_col).value = new_step_value
            modified_count += 1
        
        # 处理预期结果列（包括空单元格）
        result_value = get_cell_value(ws, row, expected_result_col)
        new_result_value = add_hash_prefix(result_value)
        # 如果单元格为空或值发生变化，则更新
        if result_value is None or new_result_value != result_value:
            ws.cell(row=row, column=expected_result_col).value = new_result_value
            modified_count += 1
    
    print(f"已为 {modified_count} 个单元格添加#号前缀")
    
    # 第二步：找到用例名称有值的行，并合并内容
    print("\n步骤2: 合并连续行的步骤描述和预期结果...")
    
    # 找到所有用例名称有值的行
    case_name_rows = []
    for row in range(2, ws.max_row + 1):
        case_name_value = get_cell_value(ws, row, case_name_col)
        if case_name_value:
            case_name_rows.append(row)
    
    print(f"找到 {len(case_name_rows)} 个用例名称有值的行: {case_name_rows}")
    
    # 合并内容
    merged_count = 0
    for i in range(len(case_name_rows)):
        current_row = case_name_rows[i]
        # 如果是最后一个用例名称行，合并到文档的最后一行
        if i + 1 < len(case_name_rows):
            next_row = case_name_rows[i + 1]
        else:
            # 最后一个用例名称行，应该合并到文档的最后一行
            next_row = ws.max_row + 1
        
        # 合并范围：从current_row到next_row-1
        # 但current_row本身的内容要保留，所以合并的是current_row+1到next_row-1的内容到current_row
        
        # 收集步骤描述
        step_descriptions = []
        # 先收集当前行的步骤描述（如果有）
        current_step = get_cell_value(ws, current_row, step_desc_col)
        if current_step:
            step_descriptions.append(current_step)
        
        # 收集后续行的步骤描述
        for row in range(current_row + 1, next_row):
            step_value = get_cell_value(ws, row, step_desc_col)
            if step_value:
                step_descriptions.append(step_value)
        
        # 收集预期结果
        expected_results = []
        # 先收集当前行的预期结果（如果有）
        current_result = get_cell_value(ws, current_row, expected_result_col)
        if current_result:
            expected_results.append(current_result)
        
        # 收集后续行的预期结果
        for row in range(current_row + 1, next_row):
            result_value = get_cell_value(ws, row, expected_result_col)
            if result_value:
                expected_results.append(result_value)
        
        # 合并到当前行（直接拼接，无分隔符）
        if step_descriptions:
            merged_step = ''.join(step_descriptions)
            ws.cell(row=current_row, column=step_desc_col).value = merged_step
        
        if expected_results:
            merged_result = ''.join(expected_results)
            ws.cell(row=current_row, column=expected_result_col).value = merged_result
        
        # 清空被合并行的步骤描述和预期结果列
        for row in range(current_row + 1, next_row):
            ws.cell(row=row, column=step_desc_col).value = None
            ws.cell(row=row, column=expected_result_col).value = None
        
        if len(step_descriptions) > 1 or len(expected_results) > 1:
            merged_count += 1
            cleared_rows = next_row - current_row - 1
            print(f"  行{current_row}: 合并了 {len(step_descriptions)} 个步骤描述, {len(expected_results)} 个预期结果，清空了 {cleared_rows} 行（行{current_row+1}到行{next_row-1}）")
    
    print(f"共合并了 {merged_count} 组内容")
    
    # 第三步：清理全部的空行
    print("\n步骤3: 清理全部的空行...")
    initial_max_row = ws.max_row
    deleted_count = 0
    
    # 从后往前遍历，避免删除后索引变化的问题
    # 从最后一行开始，到第2行结束（第1行是表头，不删除）
    # 记录初始最大行数，避免死循环
    start_row = ws.max_row
    
    for row in range(start_row, 1, -1):  # 从最后一行往前到第2行
        # 检查整行是否为空
        is_empty = True
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                # 检查是否为空字符串或只包含空白字符
                if isinstance(cell_value, str):
                    if cell_value.strip():
                        is_empty = False
                        break
                else:
                    is_empty = False
                    break
        
        # 如果整行为空，删除该行
        if is_empty:
            ws.delete_rows(row)
            deleted_count += 1
    
    print(f"删除了 {deleted_count} 个空行（初始行数: {initial_max_row}, 最终行数: {ws.max_row}）")
    
    # 保存文件
    try:
        wb.save(output_file)
        print(f"\n✅ 处理完成！文件已保存到: {output_file}")
    except Exception as e:
        print(f"\n❌ 保存文件时出错: {e}")
        sys.exit(1)
    finally:
        wb.close()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python merge_cells.py <input_file> [<output_file>]")
        print("  input_file: 输入Excel文件路径")
        print("  output_file: 输出Excel文件路径（可选，默认覆盖原文件）")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    merge_cells(input_file, output_file)
